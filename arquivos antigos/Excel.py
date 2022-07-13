from openpyxl import Workbook
from bancoDados import SelectBD
from datetime import date
from datetime import datetime
from dateutil.relativedelta import relativedelta
from OitoMinutos import gravar08minutos
from DezMinutos import gravar10minutos
from DozeMinutos import gravar12minutos

# Verifica se é ambas marcam
def isAmbasMarcam(placar: str):
    placar = placar.replace('-', '').replace('+', '')
    if int(str(placar)[0]) != 0 and int(str(placar)[1]) != 0:
        return True
    else:
        return False

def isCasa_Vence(placar: str):
    placar = placar.replace('-', '').replace(' ', '').replace('+', '')
    if int(str(placar)[0]) > int(str(placar)[1]):
        return True
    else:
        return False

def isVisitante_Vence(placar: str):
    if placar != '':
        pass
    else:
        return

    placar = placar.replace('-', '').replace(' ', '').replace('+', '')
    if int(str(placar)[0]) < int(str(placar)[1]):
        return True
    else:
        return False

# Pega as médias
def getSomaResultado(resultado : str):
    placar = str(resultado).replace('-','').replace('+','')
    try:
        return int(placar[0]) + int(placar[1])
    except:
        return 0

def getMediaGeral(jogadorA : str, jogadorB : str):

    dados = SelectBD().selectJogos(jogadorA, jogadorB)
    totalJogos = len(dados)

    print('---------------')
    print(f'total Jogos: {totalJogos}')
    
    soma = 0
    for dado in dados:
        soma += int(getSomaResultado(dado[6]))

    print(f'Soma dos gols :{soma}')
    try:
        return round(soma/totalJogos,2)
    except:
        return 0

def getMediaMesmoTime(dados : list):

    totalJogos = len(dados)
    soma = 0
    for dado in dados:
        soma += int(getSomaResultado(dado[6]))

    try:
        return round(soma / totalJogos, 2)
    except:
        return 0

def getMediaAmbasMarcam(dados : list):

    totalJogos = len(dados)
    soma = 0
    for dado in dados:
        if isAmbasMarcam(dado[6]):
            soma += 1

    try:
        return round(soma / totalJogos, 2)
    except:
        return 0

def getMediaCasa(dados : list):

    totalJogos = len(dados)
    soma = 0
    for dado in dados:
        if isCasa_Vence(dado[6]):
            soma += 1
    try:
        return round(soma / totalJogos, 2)
    except:
        return 0

def getMediaVisitante(dados : list):

    totalJogos = len(dados)
    soma = 0
    for dado in dados:
        if isVisitante_Vence(dado[6]):
            soma += 1
    try:
        return round(soma / totalJogos, 2)
    except:
        return 0

# Pega a hora de notificar
def getHoraNotificar(dataPesquisa: str, qtdDias : int):
    dataPesquisa = datetime.strptime(dataPesquisa, '%d/%m/%Y')
    horaNotificar = datetime.strftime(dataPesquisa + relativedelta(days=-qtdDias), '%d/%m/%Y')
    return horaNotificar

def getMediaSeteDias(jogadorA : str, jogadorB : str):
    dataAtual = str(date.today().strftime('%d/%m/%Y'))

    dados = []
    dia = 7
    while dia >= 1:
        data = getHoraNotificar(dataAtual, dia)
        dados += SelectBD().selectJogosPorData(jogadorA, jogadorB, data)
        dia -= 1

    totalJogos = len(dados)

    soma = 0
    for linha in dados:
        soma += int(getSomaResultado(linha[6]))

    try:
        return round(soma / totalJogos, 2)
    except:
        return 0

def getMediaDezDias(jogadorA : str, jogadorB : str):
    dataAtual = str(date.today().strftime('%d/%m/%Y'))

    dados = []
    dia = 10
    while dia >= 1:
        data = getHoraNotificar(dataAtual, dia)
        dados += SelectBD().selectJogosPorData(jogadorA, jogadorB, data)
        dia -= 1

    totalJogos = len(dados)

    soma = 0
    for linha in dados:
        soma += int(getSomaResultado(linha[6]))

    try:
        return round(soma / totalJogos, 2)
    except:
        return 0

# Cria a pasta de trabalho
wb = Workbook()
ws = wb.active
ws.title = "Dados"

# Nome das colunas
ws.cell(1, 1, 'Liga')
ws.cell(1, 2, 'jogadorA')
ws.cell(1, 3, 'jogadorB')
ws.cell(1, 4, 'timeA')
ws.cell(1, 5, 'timeB')
ws.cell(1, 6, 'data')
ws.cell(1, 7, 'hora')
ws.cell(1, 8, 'Média de Gols Geral')
ws.cell(1, 9,' Média ultimos 7 dias')
ws.cell(1, 10, 'Média ultimos 10 dias')
ws.cell(1, 11, 'Média com esses times')
ws.cell(1, 12, '% Ambos Marcam')
ws.cell(1, 13, '% Vitória Casa')
ws.cell(1, 14, '% Vitória Visitante')

data = input('Informe a data : ')
diretorioPlanilha = input('Informe onde deseja salvar a planilha : ')

# grava os placares no banco
gravar08minutos(data)
gravar10minutos(data)
gravar12minutos(data)

dadosFuturos = SelectBD().selectDadosPeriodo(data)

Linha = 2
for linha in dadosFuturos:
    ws.cell(Linha, 1, linha[6])
    ws.cell(Linha, 2, linha[1])
    ws.cell(Linha, 3, linha[2])
    ws.cell(Linha, 4, linha[3])
    ws.cell(Linha, 5, linha[4])
    ws.cell(Linha, 6, str(linha[5])[0:10])
    ws.cell(Linha, 7, str(linha[5])[11:16])
    ws.cell(Linha, 8, getMediaGeral(linha[1], linha[2]))
    ws.cell(Linha, 9, getMediaSeteDias(linha[1], linha[2]))
    ws.cell(Linha, 10, getMediaDezDias(linha[1], linha[2]))

    # dados coletados
    lin = 2
    dadosVerify = SelectBD().selectJogosPorTime(linha[1], linha[3], linha[2], linha[4])
    ws.cell(Linha, 11, getMediaMesmoTime(dadosVerify))
    ws.cell(Linha, 12, f'{getMediaAmbasMarcam(dadosVerify) * 100}%')
    ws.cell(Linha, 13, f'{getMediaCasa(dadosVerify) * 100}%')
    ws.cell(Linha, 14, f'{getMediaVisitante(dadosVerify) * 100}%')

    print(f'Gravando Linha {Linha}')
    Linha += 1

# Salva a pasta de trabalho
wb.save(diretorioPlanilha)